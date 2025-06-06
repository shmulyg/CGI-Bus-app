import pandas as pd
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os

# === Settings ===
INPUT_EXCEL = "students.xlsx"
SHEET_NAME = "Students"
BARCODE_DIR = "barcodes"

try:
    # Load Excel using pandas
    df = pd.read_excel(INPUT_EXCEL, sheet_name=SHEET_NAME)

    # Generate barcode images
    os.makedirs(BARCODE_DIR, exist_ok=True)

    for _, row in df.iterrows():
        student_id = str(row['id'])
        filename = os.path.join(BARCODE_DIR, student_id)
        code = barcode.get("code128", student_id, writer=ImageWriter())
        saved_path = code.save(filename)

        if not os.path.exists(saved_path):
            raise FileNotFoundError(f"Barcode not saved: {saved_path}")

    # Load Excel workbook with openpyxl to write images
    wb = load_workbook(INPUT_EXCEL)
    ws = wb[SHEET_NAME]

    # Add header for barcode column
    ws.cell(row=1, column=4, value="Barcode")

    # Insert images
    for i, row in enumerate(df.itertuples(), start=2):
        student_id = str(row.id)
        image_path = os.path.join(BARCODE_DIR, f"{student_id}.png")
        img = XLImage(image_path)
        img.width = 150
        img.height = 50
        ws.add_image(img, f"D{i}")

    # Save workbook
    wb.save(INPUT_EXCEL)
    print(f"✅ Barcodes added directly to {INPUT_EXCEL}")

except Exception as e:
    print(f"❌ Error: {e}")
